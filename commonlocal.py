#!/usr/bin/env python3
# Common Localizations
# This file is released into the public domain.
# The localizations are licensed under the MIT License, from http://goo.gl/fhnw1t
__doc__ = """Common Localizations.

Creates localization files from the `Common Localization project <http://goo.gl/fhnw1t>`_.

Usage:
  commonlocal.py update
  commonlocal.py generate [options]
  commonlocal.py -h | --help

Options:
  --output-folder=<string>    Folder to dump output files in [default: output]
"""
import os
import shutil
from docopt import docopt

# configuration
source_filename = 'source.xlsx'
source_url = 'https://docs.google.com/spreadsheets/d/135HgMYcRDt6vnJN0d-xFMEZUeWjVbc61ETf8uVwHERE/export?format=xlsx&id=135HgMYcRDt6vnJN0d-xFMEZUeWjVbc61ETf8uVwHERE'
source_niceurl = 'http://goo.gl/fhnw1t'
base_data_folder = 'base'

language_sheet_name = 'lang'

# helper funcs
def cell_value(cell):
    """Returns cell value or an empty string."""
    value = getattr(cell, 'value', '')
    if value is None or str(value).startswith('='):  # we don't calculate expressions
        return ''
    else:
        return str(value)

# doing stuff
if __name__ == '__main__':
    args = docopt(__doc__)

    if args['update']:
        print('Updating source file')
        import requests

        with open(source_filename, 'wb') as source_file:
            r = requests.get(source_url)
            for chunk in r.iter_content(chunk_size=512 * 1024):
                if chunk: # filter out keep-alive new chunks
                    source_file.write(chunk)
            r.close


    elif args['generate']:
        print('Generating output files')
        from openpyxl import load_workbook
        wb = load_workbook(source_filename)

        # wipe output folder
        output_folder = args['--output-folder']
        if os.path.exists(output_folder):
            shutil.rmtree(output_folder)
        os.makedirs(output_folder)

        with open(os.path.join(output_folder, 'read me'), 'w') as read_me:
            read_me.write('Do not edit the files in this folder.\n')
            read_me.write('This folder is automatically deleted and recreated each run.\n\n')
            read_me.write('Edit the source file instead at: {}\n'.format(source_niceurl))

        # base language information
        languages = {}
        language_rows = []
        language_sheet = wb.get_sheet_by_name(name=language_sheet_name)

        for row_id, row in enumerate(language_sheet.iter_rows()):
            # language names
            if row_id == 0:
                for cell_id, cell in enumerate(row):
                    if cell_id:  # skip first col here, is useless
                        lang_id = cell_value(cell).strip()
                        if lang_id:
                            language_rows.append(lang_id)
                            languages[lang_id] = {}

            # information
            else:
                key = None
                for cell_id, cell in enumerate(row):
                    if cell_id == 0:  # key
                        key = cell_value(cell).strip().replace(' ', '').replace(':', '')
                    else:
                        if key:  # if row is not empty
                            if cell_id <= len(language_rows):
                                lang_id = language_rows[cell_id - 1]
                                value = cell_value(cell).strip()
                                languages[lang_id][key] = value

        # actual translation tables
        translations = {}

        for sheet_name in wb.get_sheet_names():
            translations[sheet_name] = {}

            language_rows = []
            this_sheet = wb.get_sheet_by_name(name=sheet_name)

            for row_id, row in enumerate(this_sheet.iter_rows()):
                # language names
                if row_id == 0:
                    for cell_id, cell in enumerate(row):
                        if cell_id:  # skip first col here, is useless
                            lang_id = cell_value(cell).strip()
                            if lang_id:
                                language_rows.append(lang_id)
                                translations[sheet_name][lang_id] = {}

                # information
                else:
                    key = None
                    for cell_id, cell in enumerate(row):
                        if cell_id == 0:  # key
                            key = cell_value(cell).strip().replace(' ', '').replace(':', '')
                        else:
                            if key:  # if row is not empty
                                if cell_id <= len(language_rows):
                                    lang_id = language_rows[cell_id - 1]
                                    value = cell_value(cell).strip()
                                    translations[sheet_name][lang_id][key] = value

        # ini
        print('Generating .ini files')
        ini_folder = os.path.join(output_folder, 'ini')
        os.makedirs(ini_folder)

        for sheet_id, sheet_name in enumerate(sorted(translations)):
            for lang_id in translations[sheet_name]:
                ini_filename = os.path.join(ini_folder, '{}.ini'.format(lang_id))
                with open(ini_filename, 'a') as ini_file:
                    # separator
                    if sheet_id:
                        ini_file.write('\n')

                    # header
                    ini_file.write('[{}]\n'.format(sheet_name.replace('[', '').replace(']', '')))

                    # columns
                    for key in sorted(translations[sheet_name][lang_id]):
                        value = translations[sheet_name][lang_id][key]
                        ini_file.write('{}={}\n'.format(key, value))

        # lua
        print('Generating lua library')
        lua_folder = os.path.join(output_folder, 'lua')
        os.makedirs(lua_folder)

        lua_language_folder = os.path.join(output_folder, 'lua', 'languages')
        os.makedirs(lua_language_folder)

        lua_data_folder = os.path.join(base_data_folder, 'lua')

        # overall lua file
        lua_data_library_filename = os.path.join(lua_data_folder, 'commonlocal.lua')
        lua_data_library = open(lua_data_library_filename, 'r').read()
        nice_language_list = str(language_rows).replace('[', '{').replace(']', '}')
        lua_data_library = lua_data_library.replace('%%SUPPORTED_LANGUAGES%%', nice_language_list)

        # generate language info
        nice_language_info = 'base_language_info = {}\n'
        for lang_id in sorted(language_rows):
            nice_language_info += 'base_language_info["{lang_id}"] = {{}}\n'.format(lang_id=lang_id)
            for key in sorted(languages[lang_id]):
                nice_language_info += 'base_language_info["{lang_id}"]["{key}"] = {value}\n'.format(lang_id=lang_id, key=key, value=str([languages[lang_id][key]])[1:-1])
        lua_data_library = lua_data_library.replace('%%BASE_LANGUAGE_INFO%%', nice_language_info)

        lua_filename = os.path.join(lua_folder, 'commonlocal.lua')
        with open(lua_filename, 'w') as lua_file:
            lua_file.write(lua_data_library)

        # specific language files
        lua_data_format_filename = os.path.join(lua_data_folder, 'lang_format.lua')
        lua_data_format = open(lua_data_format_filename, 'r').read()

        for lang_id in language_rows:
            lua_filename = os.path.join(lua_language_folder, '{}.lua'.format(lang_id))

            lang_name = languages[lang_id]['nameen']
            lang_dict = {}

            for sheet_name in sorted(translations):
                lang_dict[sheet_name] = translations[sheet_name].get(lang_id, {})

            with open(lua_filename, 'w') as lua_file:
                nice_lang_dict = 'lang_dict = {}\n'
                for section in sorted(lang_dict):
                    nice_lang_dict += 'lang_dict["{section}"] = {{}}\n'.format(section=section)
                    for key in sorted(lang_dict[section]):
                        value = lang_dict[section][key].replace('\\', '\\\\').replace('"', '\\"')
                        nice_lang_dict += 'lang_dict["{section}"]["{key}"] = "{value}"\n'.format(section=section, key=key, value=value)
                lua_file.write(lua_data_format.format(lang_name=lang_name, lang_dict=nice_lang_dict))
